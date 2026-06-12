# -*- coding: utf-8 -*-
"""Excel 多工作表輸出：主表、四個子榜、參數與資料來源說明。"""

import io
import datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

# 主表 24 欄：基本資訊 ＋ 免費替代指標（皆為公開資料機械式計算，非投資建議）
COLUMNS = [
    "股票代號", "股票名稱", "所屬產業", "目前股價", "成交量(張)",
    "本益比", "ROE", "負債比",
    "近四季EPS", "EPS年增率", "近3月營收YoY", "近6月營收YoY", "近12月營收YoY",
    "營收是否加速", "估算合理價", "估算上漲空間", "PEG替代值",
    "法人近20日買賣超(張)", "下半年成長題材", "主要風險",
    "建議買進區間", "停損區間", "評分", "投資評等",
]

_HEAD_FILL = PatternFill("solid", fgColor="1F3864")
_HEAD_FONT = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=10)
_ESTIMATE_FILL = PatternFill("solid", fgColor="FFF2CC")  # 機械式估算欄位淡黃底
_CELL_FONT = Font(name="Microsoft JhengHei", size=10)
_TITLE_FONT = Font(name="Microsoft JhengHei", bold=True, size=13, color="1F3864")
_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
# 標示為「機械式估算」的欄位（淡黃底提醒：非券商目標價／法人預估）
_ESTIMATE_COLS = {"估算合理價", "估算上漲空間", "PEG替代值"}


def _write_table(ws, df: pd.DataFrame, title: str):
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    start = 3
    for j, col in enumerate(COLUMNS, start=1):
        c = ws.cell(row=start, column=j, value=col)
        c.fill = _HEAD_FILL
        c.font = _HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
    for i, (_, row) in enumerate(df.iterrows(), start=start + 1):
        for j, col in enumerate(COLUMNS, start=1):
            val = row.get(col, "")
            c = ws.cell(row=i, column=j, value=val)
            c.font = _CELL_FONT
            c.border = _BORDER
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col in _ESTIMATE_COLS:
                c.fill = _ESTIMATE_FILL
    widths = [9, 11, 13, 9, 9, 7, 7, 7, 9, 9, 10, 10, 10, 9, 10, 11, 9,
              16, 20, 24, 18, 16, 7, 8]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=start + 1, column=1)


def _format_counts(counts: dict, limit: int = 8):
    if not counts:
        return ["  無"]
    items = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))[:limit]
    lines = [f"  {name}：{count} 檔" for name, count in items]
    if len(counts) > limit:
        lines.append(f"  其他：{sum(counts.values()) - sum(count for _, count in items)} 檔")
    return lines


def _write_notes(ws, params_used: dict, total_passed: int, stats: dict = None):
    ws["A1"] = "篩選參數與資料來源說明"
    ws["A1"].font = _TITLE_FONT
    stats = stats or {}
    lines = [
        "",
        f"產出時間：{dt.datetime.now():%Y-%m-%d %H:%M}",
        f"通過全部基本面硬門檻的個股數：{total_passed}",
        f"最終入選並評分的個股數：{stats.get('selected_count', total_passed)}",
        "",
        "【本次資料品質摘要】",
        f"  全市場有效個股：{stats.get('market_count', '未記錄')}",
        f"  進入逐檔深掃：{stats.get('stage1_count', '未記錄')}",
        f"  通過技術面與位階篩選：{stats.get('technical_pass_count', '未記錄')}",
        f"  通過基本面硬門檻：{stats.get('fundamentals_pass_count', total_passed)}",
        "  略過原因：",
        *_format_counts(stats.get("skipped", {})),
        "  錯誤原因：",
        *_format_counts(stats.get("errors", {})),
        "",
        "【資料來源】（皆為免費公開資料）",
        "  價量/估值：證券交易所 OpenAPI、櫃買中心 OpenAPI（全市場批次）",
        "  基本面/法人：FinMind 開放資料（逐檔）",
        "",
        "【免費替代指標說明・本表以淡黃底標記的估算欄位】",
        "  本系統不使用任何付費預估資料（不抓券商共識 EPS／目標價／法人預估）。",
        "  EPS 年增率、近3/6/12月營收 YoY、估算合理價、PEG 替代值，",
        "  皆由公開歷史資料（FinMind 財報與月營收）機械式計算。",
        "  ・估算合理價 = 近四季EPS × 同產業 PE 中位數（不是券商目標價）。",
        "  ・估算上漲空間 = 估算合理價 ÷ 現價 − 1。",
        "  ・PEG 替代值 = 本益比 ÷ EPS年增率(%)（不是法人預估 PEG）。",
        "  以上僅供篩選參考，資料不足者顯示「資料不足」，非投資建議。",
        "",
        "【套用門檻】",
        f"  最近四季單季 EPS 皆為正：{params_used['eps']}",
        f"  近一年(TTM)營收年增率 > {params_used['rev']:.0%}",
        f"  最新單季毛利率衰退 ≤ {params_used['gm']} 個百分點",
        f"  ROE > {params_used['roe']:.0%}",
        f"  負債比 < {params_used['debt']:.0%}",
        f"  近 {params_used['noloss']} 年無年度虧損",
        f"  日成交量 > {params_used['lots']:,} 張、距 52 週高 ≤ {params_used['dist']:.0%}",
        f"  PE 必須低於產業基準：{params_used.get('pe_below', '否')}",
        f"  近5日融資餘額增幅警示門檻：>{params_used.get('margin_surge', 0):.0%}"
        f"（排除：{params_used.get('exclude_margin_surge', '否')}）",
        f"  停損參考：季線下方 {params_used.get('stop_buffer', 0):.0%}",
        f"  排除跌破年線、長空排列",
        "",
        "【評等說明】",
        "  「投資評等」為各客觀篩選條件的符合度評分（A+/A/B），",
        "  「建議買進區間／停損區間」為移動平均線推導之技術參考位階。",
        "  以上皆為機械式計算結果，非投資建議，使用者應自行判斷並承擔風險。",
        "",
        "【免責聲明】",
        "  本表僅供研究與資料整理，不構成任何投資建議或要約，提供者非投資顧問。",
        "  資料可能有誤差或延遲，實際交易請以官方公告與券商資訊為準。",
    ]
    for i, line in enumerate(lines, start=2):
        c = ws.cell(row=i, column=1, value=line)
        c.font = _CELL_FONT
    ws.column_dimensions["A"].width = 80


def build_workbook(main_df, sublists: dict, params_used: dict, total_passed: int,
                   stats: dict = None, cfg=None):
    cfg = cfg or config
    wb = Workbook()
    ws = wb.active
    ws.title = "前20名逢低布局"
    _write_table(ws, main_df.head(cfg.TOP_N), f"恐慌殺盤逢低布局・前 {cfg.TOP_N} 名")

    for name, df in sublists.items():
        w = wb.create_sheet(name)
        _write_table(w, df, name)

    _write_notes(wb.create_sheet("參數與免責"), params_used, total_passed, stats)
    return wb


def write_report(main_df, sublists: dict, params_used: dict, total_passed: int, stats: dict = None, cfg=None):
    wb = build_workbook(main_df, sublists, params_used, total_passed, stats=stats, cfg=cfg)
    path = f"{config.OUTPUT_DIR}/逢低布局選股_{dt.datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(path)
    return path


def build_report_bytes(main_df, sublists: dict, params_used: dict, total_passed: int,
                       stats: dict = None, cfg=None) -> bytes:
    wb = build_workbook(main_df, sublists, params_used, total_passed, stats=stats, cfg=cfg)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
